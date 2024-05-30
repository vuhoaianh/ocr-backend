from openpyxl import Workbook
import json


def table2excel(data, filename="table.xlsx"):
    metadata = data["metadata"]
    wb = Workbook()
    sheet = wb.active

    table_metadata = metadata["table_metadata"][0]
    table_data= table_metadata["table_data"]

    rows = table_data["rows"]
    row_id = 0
    for row in rows:
        for col, cells in enumerate(row):
            if not cells:
                txt = ''
                c = sheet.cell(row=row_id, column=col+1)
                c.value = txt
            if len(cells) == 1:
                cell = cells[0]
                txt = cell['text']
                c = sheet.cell(row=row_id+1, column=col+1)
                c.value = txt
            else:
                for cell in cells:
                    txt = cell['text']
                    c = sheet.cell(row=row_id+1, column=col+1)
                    c.value = txt
                    row_id += 1
        row_id += 1
    wb.save(filename)